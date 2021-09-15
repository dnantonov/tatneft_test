import json
from flask import Flask, jsonify
from flask import request
from openpyxl import load_workbook
from flasgger import Swagger

flask_app = Flask(__name__)
swagger = Swagger(flask_app)


@flask_app.route("/test/service/<int:sheet_id>/<sheet_type>/", methods=["GET"])
def get_data(sheet_id, sheet_type):
    """Микросервис "Тестовый сервис"
        ---
        parameters:
            - in: path
              name: sheet_id
              type: integer
              required: true
        responses:
          200:
            description: Get data
        """
    wb = load_workbook('тест.xlsx')
    sheet = wb['Лист1']
    with open('тест.json', encoding='utf-8') as json_file:
        json_data = json.load(json_file)
    try:
        data = {
            'id': sheet[f'A{sheet_id + 1}'].value,
            'name': sheet[f'B{sheet_id + 1}'].value,
            'data': sheet[f'C{sheet_id + 1}'].value,
        }
        sheet_type = json_data[data['id']]
        data['type'] = sheet_type
    except KeyError:
        return '<h2>Object Does Not Exist</h2>', 404
    return jsonify(data)


@flask_app.route("/test/service/", methods=['POST'])
def post_data():
    """Микросервис "Тестовый сервис"
            post endpoint
            ---
            parameters:
              - name: data
                type: string
                in: body
                required: true
            responses:
              200:
                description: Update data
            """
    content = request.get_json()
    wb = load_workbook('тест.xlsx')
    sheet = wb['Лист1']
    sheet_id = content['id'] + 1
    if sheet[f'B{sheet_id}'].value is None:
        return '<h2>Object Does Not Exist</h2>', 404
    sheet[f'B{sheet_id}'].value = content['name']
    wb.save(filename="тест.xlsx")
    return jsonify({"info": "ok"})


if __name__ == '__main__':
    flask_app.run(debug=True)
